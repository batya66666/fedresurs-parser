import time
import random
import re
import logging
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional
import os
from openpyxl import load_workbook
import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# =========================
# LOGGING
# =========================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)
log = logging.getLogger("FedresursParser")


# =========================
# MODELS (1-в-1 с Java)
# =========================
@dataclass
class LegalEntity:
    fullName: str = ""
    inn: str = ""
    ogrn: str = ""
    kpp: str = ""
    authorizedCapital: str = ""
    registrationDate: str = ""
    address: str = ""
    region: str = ""
    legalForm: str = ""
    okved: str = ""
    status: str = ""
    procedureType: str = ""
    caseNumber: str = ""
    caseStatus: str = ""
    caseEndDate: str = ""
    arbitrationManagerName: str = ""
    arbitrationManagerInn: str = ""
    managerAppointmentDate: str = ""
    publicationsCount: str = "0"
    tradesCount: str = "0"
    sourceUrl: str = ""


@dataclass
class PhysicalPerson:
    fullName: str = ""
    previousFullName: str = ""
    inn: str = ""
    snils: str = ""
    birthDate: str = ""
    birthPlace: str = ""
    residenceAddress: str = ""
    region: str = ""
    entrepreneurOgrnip: str = ""
    entrepreneurStatus: str = ""
    okved: str = ""
    registrationDate: str = ""
    terminationDate: str = ""
    bankruptcyStatus: str = ""
    procedureType: str = ""
    caseNumber: str = ""
    arbitrationManagerName: str = ""
    sourceUrl: str = ""


# =========================
# UTILS (как v() в Java + даты)
# =========================
def v(val: Any) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if not s:
        return ""
    low = s.lower()
    if low in {"null", "н/д", "н.д.", "нет"}:
        return ""
    if s == "-" :
        return ""
    return s


def format_date(date_str: Any) -> str:
    if date_str is None:
        return ""
    s = str(date_str).strip()
    if not s or s.lower() == "null":
        return ""
    try:
        clean = s.split("T")[0] if "T" in s else s
        dt = datetime.strptime(clean, "%Y-%m-%d")
        return dt.strftime("%d.%m.%Y")
    except Exception:
        # как в Java: если не смогли распарсить — вернём как есть
        return s


def jpath(obj: Any, *keys: str, default: str = "") -> str:
    """
    Безопасный доступ как node.path("a").path("b").asText("")
    """
    cur = obj
    for k in keys:
        if not isinstance(cur, dict):
            return default
        cur = cur.get(k)
    if cur is None:
        return default
    return str(cur)


# =========================
# API SERVICE (1-в-1 логика Java)
# =========================
class ApiService:
    COMPANY_LIST_URL = "https://bankrot.fedresurs.ru/backend/cmpbankrupts"
    PERSON_LIST_URL  = "https://bankrot.fedresurs.ru/backend/prsnbankrupts"

    MAX_ATTEMPTS = 4
    BASE_BACKOFF_MS = 700

    def __init__(self):
        self.s = requests.Session()
        self.s.headers.update({
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
            "Accept": "application/json, text/plain, */*",
        })
        self.timeout = 30

    def send_get_with_retry(self, url: str, referer: str) -> Dict[str, Any]:
        backoff = self.BASE_BACKOFF_MS / 1000.0
        for attempt in range(1, self.MAX_ATTEMPTS + 1):
            try:
                r = self.s.get(url, headers={"Referer": referer}, timeout=self.timeout)
                code = r.status_code

                if code == 200:
                    # иногда могут вернуть не-json: подстрахуемся
                    try:
                        return r.json()
                    except Exception:
                        return {}

                retryable = (code == 429) or (500 <= code <= 599)
                if not retryable:
                    return {}

                jitter = random.uniform(0, 0.25)
                time.sleep(backoff + jitter)
                backoff *= 2
            except Exception:
                # сетевые ошибки тоже ретраим
                jitter = random.uniform(0, 0.25)
                time.sleep(backoff + jitter)
                backoff *= 2

        return {}

    def fetch_json_with_retry(self, url: str, referer: str) -> Optional[Dict[str, Any]]:
        node = self.send_get_with_retry(url, referer)
        if not node:
            return None
        return node

    def fetch_list_items(self, is_legal: bool, offset: int, limit: int) -> List[Dict[str, Any]]:
        base = self.COMPANY_LIST_URL if is_legal else self.PERSON_LIST_URL
        url = f"{base}?limit={limit}&offset={offset}"
        root = self.send_get_with_retry(url, "https://bankrot.fedresurs.ru/bankrupts")

        page_data = root.get("pageData")
        if isinstance(page_data, list):
            return page_data
        return []

    def count_biddings(self, guid: str, referer: str) -> int:
        total, offset, limit = 0, 0, 50
        try:
            while True:
                url = (
                    "https://fedresurs.ru/backend/biddings"
                    f"?limit={limit}&offset={offset}&bankruptGuid={guid}"
                )
                resp = self.fetch_json_with_retry(url, referer)
                if not resp:
                    break
                arr = resp.get("pageData")
                if not isinstance(arr, list) or len(arr) == 0:
                    break

                total += len(arr)
                if len(arr) < limit:
                    break
                offset += limit
        except Exception:
            return 0
        return total

    def _extract_company_name(self, full_name: str) -> str:
        if not full_name:
            return ""
        m = re.search(r'«(.+)»', full_name)
        if m:
            return m.group(1)
        start = full_name.find('"')
        end = full_name.rfind('"')
        if start != -1 and end > start:
            return full_name[start+1:end]
        return full_name

    def fetch_legal_full_details(self, list_item: Dict[str, Any]) -> LegalEntity:
        guid = v(list_item.get("guid"))
        e = LegalEntity()

        lc = list_item.get("lastLegalCase") or {}
        e.caseNumber = v(jpath(lc, "number"))
        e.arbitrationManagerName = v(jpath(lc, "arbitrManagerFio"))

        status_desc = v(jpath(lc, "status", "description"))
        status_code = v(jpath(lc, "status", "code"))
        status_date = v(jpath(lc, "status", "date"))

        e.status = status_desc
        e.procedureType = status_code
        e.caseStatus = "Завершено" if "завершено" in status_desc.lower() else "Активно"
        e.caseEndDate = format_date(status_date)

        e.inn = v(list_item.get("inn"))
        e.ogrn = v(list_item.get("ogrn"))
        e.region = v(list_item.get("region"))
        e.sourceUrl = f"https://fedresurs.ru/company/{guid}"

        try:
            details = self.fetch_json_with_retry(f"https://fedresurs.ru/backend/companies/{guid}", e.sourceUrl)
            if details:
                raw_name = v(details.get("fullName")) or v(list_item.get("name"))
                e.fullName = self._extract_company_name(raw_name)
                e.kpp = v(details.get("kpp"))
                e.address = v(details.get("addressEgrul"))
                e.authorizedCapital = v(details.get("authorizedCapital"))

                okopf = details.get("okopf") or {}
                okved = details.get("okved") or {}
                e.legalForm = v(okopf.get("name"))
                e.okved = v(okved.get("name"))

                e.registrationDate = format_date(details.get("dateReg"))

            ieb = self.fetch_json_with_retry(f"https://fedresurs.ru/backend/companies/{guid}/ieb", e.sourceUrl)
            if ieb and isinstance(ieb.get("pageData"), list) and len(ieb["pageData"]) > 0:
                m = ieb["pageData"][0]
                e.arbitrationManagerInn = v(m.get("inn"))
                e.managerAppointmentDate = format_date(m.get("egrulDateCreate"))

            pubs = self.fetch_json_with_retry(
                f"https://fedresurs.ru/backend/companies/{guid}/publications?limit=1",
                e.sourceUrl
            )
            if pubs:
                e.publicationsCount = str(pubs.get("found") or 0)

            e.tradesCount = str(self.count_biddings(guid, e.sourceUrl))

        except Exception as ex:
            log.debug("Legal details failed guid=%s: %s", guid, ex)

        # чистим значения как в ExcelExporter.v()
        for field in e.__dataclass_fields__:
            setattr(e, field, v(getattr(e, field)))

        return e

    def fetch_physical_full_details(self, list_item: Dict[str, Any]) -> PhysicalPerson:
        guid = v(list_item.get("guid"))
        p = PhysicalPerson()

        # из списка
        p.fullName = v(list_item.get("fio"))
        p.inn = v(list_item.get("inn"))
        p.snils = v(list_item.get("snils"))
        p.region = v(list_item.get("region"))
        p.sourceUrl = f"https://fedresurs.ru/person/{guid}"

        lc = list_item.get("lastLegalCase") or {}
        p.caseNumber = v(jpath(lc, "number"))
        p.arbitrationManagerName = v(jpath(lc, "arbitrManagerFio"))
        p.bankruptcyStatus = v(jpath(lc, "status", "description"))
        p.procedureType = v(jpath(lc, "status", "code"))

        try:
            details = self.fetch_json_with_retry(f"https://fedresurs.ru/backend/persons/{guid}", p.sourceUrl)
            if details:
                p.birthDate = format_date(details.get("birthdateBankruptcy"))
                p.birthPlace = v(details.get("birthplaceBankruptcy"))
                p.residenceAddress = v(details.get("address"))

                history = details.get("nameHistories")
                if isinstance(history, list) and len(history) > 0:
                    p.previousFullName = ", ".join([v(x) for x in history if v(x)])

            ip_json = self.fetch_json_with_retry(
                f"https://fedresurs.ru/backend/persons/{guid}/individual-entrepreneurs?limit=50&offset=0",
                p.sourceUrl
            )

            if ip_json and isinstance(ip_json.get("pageData"), list) and len(ip_json["pageData"]) > 0:
                best = None
                for ip in ip_json["pageData"]:
                    if best is None:
                        best = ip
                        continue
                    d1 = v(ip.get("dateReg"))
                    d2 = v(best.get("dateReg"))
                    if d1 and (not d2 or d1 > d2):
                        best = ip

                if best:
                    p.entrepreneurOgrnip = v(best.get("ogrnip"))

                    st = best.get("status") or {}
                    p.entrepreneurStatus = v(st.get("name"))
                    p.terminationDate = format_date(st.get("date"))

                    ok = best.get("okved") or {}
                    p.okved = v(ok.get("name"))

                    p.registrationDate = format_date(best.get("dateReg"))

        except Exception as ex:
            log.debug("Physical details failed guid=%s: %s", guid, ex)

        # чистим значения как в ExcelExporter.v()
        for field in p.__dataclass_fields__:
            setattr(p, field, v(getattr(p, field)))

        return p


# =========================
# EXCEL EXPORTER (1-в-1 заголовки + листы)
# =========================
class ExcelExporter:
    LEGAL_COLS = [
        "Полное наименование", "ИНН", "ОГРН", "КПП", "Уставный капитал", "Дата регистрации",
        "Адрес", "Регион", "ОКОПФ", "ОКВЭД", "Статус банкротства", "Тип процедуры",
        "№ дела", "Статус дела", "Дата завершения", "ФИО управляющего",
        "ИНН управляющего", "Дата внесения в ЕГРЮЛ", "Публикации",
        "Торги", "URL"
    ]

    PHYS_COLS = [
        "ФИО", "Ранее имевшееся ФИО", "ИНН", "СНИЛС", "Дата рожд.", "Место рожд.",
        "Адрес проживания", "Регион", "ОГРНИП", "Статус ИП",
        "Вид деятельности", "Дата регистрации ИП", "Дата прекращения ИП", "Статус банкротства",
        "Тип процедуры", "№ дела", "Управляющий", "URL"
    ]

    def export_resume(self, legal_entities, physical_persons, file_name: str) -> None:
        """
        1) Если file_name существует — ДОПИСЫВАЕМ в него
        2) Если нет — создаём новый
        3) Дубли отсекаем по URL (последняя колонка)
        """
        if os.path.exists(file_name):
            wb = load_workbook(file_name)
        else:
            wb = Workbook()

        # --- Legal sheet
        if "LegalEntities" in wb.sheetnames:
            ws_legal = wb["LegalEntities"]
        else:
            ws_legal = wb.active
            ws_legal.title = "LegalEntities"
            self._init_sheet(ws_legal, self.LEGAL_COLS)

        # --- Physical sheet
        if "PhysicalPersons" in wb.sheetnames:
            ws_phys = wb["PhysicalPersons"]
        else:
            ws_phys = wb.create_sheet("PhysicalPersons")
            self._init_sheet(ws_phys, self.PHYS_COLS)

        # соберём уже существующие URL, чтобы исключать повторы
        legal_existing = self._read_existing_urls(ws_legal, url_col_idx=len(self.LEGAL_COLS))
        phys_existing  = self._read_existing_urls(ws_phys,  url_col_idx=len(self.PHYS_COLS))

        # дописываем только новые
        added_legal = self._append_legal(ws_legal, legal_entities, legal_existing)
        added_phys  = self._append_physical(ws_phys, physical_persons, phys_existing)

        # можно включить автоширину только по новым данным (или вообще убрать, чтобы быстрее было)
        self._autosize(ws_legal, len(self.LEGAL_COLS))
        self._autosize(ws_phys,  len(self.PHYS_COLS))

        wb.save(file_name)

        log.info("Excel обновлён: %s | добавлено ЮЛ=%d, ФЛ=%d | пропущено дублей ЮЛ=%d, ФЛ=%d",
                 file_name,
                 added_legal, added_phys,
                 max(0, len(legal_entities) - added_legal),
                 max(0, len(physical_persons) - added_phys))

    def _init_sheet(self, ws, cols):
        ws.append(cols)
        for i in range(1, len(cols) + 1):
            ws.cell(row=1, column=i).font = Font(bold=True)
        ws.freeze_panes = "A2"

    def _read_existing_urls(self, ws, url_col_idx: int) -> set:
        """
        url_col_idx: 1-based индекс колонки URL (у нас это последняя колонка)
        """
        urls = set()
        # пропускаем header row (1)
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=url_col_idx).value
            u = v(val)
            if u:
                urls.add(u)
        return urls

    def _append_legal(self, ws, entities, existing_urls: set) -> int:
        added = 0
        for e in entities:
            url = v(getattr(e, "sourceUrl", ""))
            if not url or url in existing_urls:
                continue

            ws.append([
                v(e.fullName),
                v(e.inn),
                v(e.ogrn),
                v(e.kpp),
                v(e.authorizedCapital),
                v(e.registrationDate),
                v(e.address),
                v(e.region),
                v(e.legalForm),
                v(e.okved),
                v(e.status),
                v(e.procedureType),
                v(e.caseNumber),
                v(e.caseStatus),
                v(e.caseEndDate),
                v(e.arbitrationManagerName),
                v(e.arbitrationManagerInn),
                v(e.managerAppointmentDate),
                v(e.publicationsCount),
                v(e.tradesCount),
                url,
            ])
            existing_urls.add(url)
            added += 1
        return added

    def _append_physical(self, ws, persons, existing_urls: set) -> int:
        added = 0
        for p in persons:
            url = v(getattr(p, "sourceUrl", ""))
            if not url or url in existing_urls:
                continue

            ws.append([
                v(p.fullName),
                v(p.previousFullName),
                v(p.inn),
                v(p.snils),
                v(p.birthDate),
                v(p.birthPlace),
                v(p.residenceAddress),
                v(p.region),
                v(p.entrepreneurOgrnip),
                v(p.entrepreneurStatus),
                v(p.okved),
                v(p.registrationDate),
                v(p.terminationDate),
                v(p.bankruptcyStatus),
                v(p.procedureType),
                v(p.caseNumber),
                v(p.arbitrationManagerName),
                url,
            ])
            existing_urls.add(url)
            added += 1
        return added

    def _autosize(self, ws, cols_count: int):
        for col in range(1, cols_count + 1):
            letter = get_column_letter(col)
            max_len = 0
            for cell in ws[letter]:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 2, 60)

# =========================
# MAIN (как FedresursParserApp)
# =========================
def main():
    log.info("====================================================")
    log.info("ЗАПУСК ПРОФЕССИОНАЛЬНОГО ПАРСЕРА (ГЛУБОКИЙ СБОР)")
    log.info("====================================================")

    TARGET_PER_TYPE = 50
    LIMIT = 15
    SLEEP_MS = 1500

    api = ApiService()
    exporter = ExcelExporter()

    legals: List[LegalEntity] = []
    physicals: List[PhysicalPerson] = []

    try:
        log.info(">>> Фаза 1: Сбор Юридических лиц через GUID карточки...")
        offsetL = 0
        while len(legals) < TARGET_PER_TYPE:
            items = api.fetch_list_items(True, offsetL, LIMIT)
            if not items:
                break

            for item in items:
                if len(legals) >= TARGET_PER_TYPE:
                    break

                le = api.fetch_legal_full_details(item)
                legals.append(le)

                log.info(f"[{len(legals)}/{TARGET_PER_TYPE}] ЮЛ: {le.fullName} (GUID найден)")
                time.sleep(SLEEP_MS / 1000.0)

            offsetL += LIMIT

        log.info(">>> Фаза 2: Сбор Физических лиц и ИП через GUID карточки...")
        offsetP = 0
        while len(physicals) < TARGET_PER_TYPE:
            items = api.fetch_list_items(False, offsetP, LIMIT)
            if not items:
                break

            for item in items:
                if len(physicals) >= TARGET_PER_TYPE:
                    break

                pp = api.fetch_physical_full_details(item)
                physicals.append(pp)

                status_info = "+ История имен" if v(pp.previousFullName) else ""
                log.info(f"[{len(physicals)}/{TARGET_PER_TYPE}] ФЛ: {pp.fullName} {status_info}".strip())
                time.sleep(SLEEP_MS / 1000.0)

            offsetP += LIMIT

        file_name = "fedresurs_deep_parsed.xlsx"
        log.info(">>> Сохранение данных в Excel: %s", file_name)
        exporter.export_resume(legals, physicals, file_name)

        log.info("ГОТОВО! Файл сохранен.")

    except Exception as e:
        log.exception("Критическая ошибка: %s", e)


if __name__ == "__main__":
    main()
