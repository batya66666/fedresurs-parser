from openpyxl import Workbook

class ExcelWriter:
    def __init__(self, filename):
        self.wb = Workbook()
        self.ws1 = self.wb.active
        self.ws1.title = "LegalEntities"
        self.ws2 = self.wb.create_sheet("PhysicalPersons")
        self.file = filename

    def write(self, sheet, rows):
        if not rows:
            return
        ws = self.ws1 if sheet=="legal" else self.ws2
        if ws.max_row == 1:
            ws.append(list(rows[0].keys()))
        for r in rows:
            ws.append(list(r.values()))

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

        self.wb.save(self.file)
